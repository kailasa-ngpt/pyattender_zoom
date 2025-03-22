from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import JSONResponse
import hmac
import hashlib
import json
from typing import Dict, List, Any, Optional, Set
import os
from datetime import datetime, timedelta
import pendulum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pathlib
import asyncio
import time
from contextlib import asynccontextmanager

def import_time() -> str:
    return datetime.utcnow().isoformat()
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Define lifespan context manager
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: This will be called when the FastAPI app starts
    asyncio.create_task(meeting_state._schedule_eod_reports())

    yield

    # Shutdown: Any cleanup needed when shutting down
    pass

# Create FastAPI app with lifespan
app = FastAPI(lifespan=lifespan)

class ZoomAccountManager:
    def __init__(self):
        self.tokens: List[str] = []
        self.verified_tokens: Dict[str, bool] = {}
        self.load_tokens_from_env()

    def load_tokens_from_env(self):
        """Load webhook secret tokens and their verification status from environment variables"""
        i = 1
        while True:
            token_entry = os.getenv(f'ZOOM_WEBHOOK_SECRET_{i}')
            if token_entry is None:
                break

            # Check if token has a verification state attached with '|' separator
            if '|' in token_entry:
                token, verified_str = token_entry.split('|', 1)
                verified = verified_str.lower() == 'true'
            else:
                token = token_entry
                verified = False

            self.tokens.append(token)
            self.verified_tokens[token] = verified
            i += 1

        if not self.tokens:
            raise ValueError("No Zoom webhook secret tokens found in environment variables")

        # Log current verification status
        verified_count = sum(1 for v in self.verified_tokens.values() if v)
        print(f"Loaded {len(self.tokens)} tokens, {verified_count} already verified")

    def save_verification_status(self):
        """Save verification status to .env file"""
        try:
            # Read existing .env file
            env_path = os.path.join(os.getcwd(), '.env')
            if not os.path.exists(env_path):
                print("Warning: .env file not found, cannot save verification status")
                return

            with open(env_path, 'r') as file:
                lines = file.readlines()

            # Update token verification status in .env content
            new_lines = []
            for line in lines:
                line = line.strip()
                if line.startswith('ZOOM_WEBHOOK_SECRET_'):
                    # Extract the token number and current value
                    parts = line.split('=', 1)
                    if len(parts) != 2:
                        new_lines.append(line)
                        continue

                    key, value = parts
                    token_num = key.split('_')[-1]

                    # Find if this is a token we know about
                    try:
                        token_index = int(token_num) - 1
                        if token_index < len(self.tokens):
                            token = self.tokens[token_index]
                            # Replace or add verification status
                            if '|' in value:
                                token_value = value.split('|', 1)[0]
                            else:
                                token_value = value
                            new_line = f"{key}={token_value}|{str(self.verified_tokens[token]).lower()}"
                            new_lines.append(new_line)
                        else:
                            new_lines.append(line)
                    except (ValueError, IndexError):
                        new_lines.append(line)
                else:
                    new_lines.append(line)

            # Write updated content back to .env file
            with open(env_path, 'w') as file:
                file.write('\n'.join(new_lines))

            print(f"Saved verification status to .env file")

        except Exception as e:
            print(f"Error saving verification status: {e}")

    def get_next_unverified_token(self) -> str:
        """Get the next unverified token in sequence"""
        for token in self.tokens:
            if not self.verified_tokens[token]:
                return token
        return self.tokens[0]  # If all verified, return first token

    def mark_token_as_verified(self, token: str):
        """Mark a specific token as verified and save status"""
        if token in self.verified_tokens:
            self.verified_tokens[token] = True
            # Save the updated verification status to .env
            self.save_verification_status()

    def verify_signature(self, signature: str, message: str) -> bool:
        """Try to verify signature with all tokens"""
        if not signature.startswith("v0="):
            return False

        received_hash = signature[3:]  # Remove 'v0='

        for token in self.tokens:
            expected_hash = generate_hash(message, token)
            if hmac.compare_digest(received_hash, expected_hash):
                return True
        return False

# Initialize the account manager
account_manager = ZoomAccountManager()

def generate_hash(message: str, secret: str) -> str:
    """Generate HMAC SHA-256 hash for a message using the secret token"""
    return hmac.new(
        secret.encode('utf-8'),
        message.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()

class MeetingStateManager:
    def __init__(self):
        self.meetings: Dict[str, Dict[str, Any]] = {}
        self.et_tz = pendulum.timezone('America/New_York')

        # Create folders for Raw and Reports
        self.raw_dir = pathlib.Path("Raw")
        self.reports_dir = pathlib.Path("Reports")
        self.raw_dir.mkdir(exist_ok=True)
        self.reports_dir.mkdir(exist_ok=True)

        # We'll start the background tasks when FastAPI starts

    async def _schedule_eod_reports(self):
        """Schedule EOD reports at 11:15 PM ET every day"""
        while True:
            now = pendulum.now(self.et_tz)
            # Schedule for 11:15 PM ET (after the day's meetings should be done)
            target = now.replace(hour=23, minute=15, second=0, microsecond=0)
            if now > target:
                target = target.add(days=1)

            # Sleep until the target time
            seconds_to_wait = (target - now).total_seconds()
            await asyncio.sleep(seconds_to_wait)

            # Generate EOD reports for all active meetings
            for meeting_uuid in list(self.meetings.keys()):
                await self.generate_eod_report(meeting_uuid)

            # Sleep a bit to avoid duplicate runs
            await asyncio.sleep(60)

    def store_raw_webhook(self, meeting_uuid: str, data: Dict[str, Any]) -> None:
        """Store raw webhook data to file system"""
        # Get current date for folder structure
        now = pendulum.now()
        date_str = now.format('YYYY-MM-DD')
        timestamp = now.format('YYYY-MM-DD_HH-mm-ss-SSS')

        # Create directory structure
        day_dir = self.raw_dir / date_str
        meeting_dir = day_dir / meeting_uuid
        meeting_dir.mkdir(parents=True, exist_ok=True)

        # Write webhook data to file
        file_path = meeting_dir / f"{timestamp}.json"
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=2)

        # Also print to console for debugging
        print(f"Webhook received: {json.dumps(data)}")

    def process_webhook(self, data: Dict[str, Any]) -> None:
        """Process a webhook and update meeting state"""
        event_type = data.get("event")

        # Extract meeting UUID from different event types
        meeting_uuid = None
        topic = None

        if "payload" in data and "object" in data["payload"]:
            obj = data["payload"]["object"]

            if "uuid" in obj:
                meeting_uuid = obj["uuid"]

            if "topic" in obj:
                topic = obj["topic"]

        # Store raw webhook data
        if meeting_uuid:
            self.store_raw_webhook(meeting_uuid, data)

        # Process based on event type
        if event_type == "meeting.started":
            self._handle_meeting_started(data)
        elif event_type == "meeting.ended":
            self._handle_meeting_ended(data)
        elif event_type == "meeting.participant_joined":
            self._handle_participant_joined(data)
        elif event_type == "meeting.participant_left":
            self._handle_participant_left(data)
        elif event_type == "meeting.chat_message_sent":
            self._handle_chat_message(data)

    def _handle_meeting_started(self, data: Dict[str, Any]) -> None:
        """Handle meeting.started webhook"""
        if "payload" not in data or "object" not in data["payload"]:
            return

        obj = data["payload"]["object"]
        meeting_uuid = obj.get("uuid")
        if not meeting_uuid:
            return

        # Initialize meeting state if not exists
        if meeting_uuid not in self.meetings:
            self.meetings[meeting_uuid] = {
                "topic": obj.get("topic", "Unnamed Meeting"),
                "start_time": obj.get("start_time"),
                "timezone": obj.get("timezone"),
                "host_id": obj.get("host_id"),
                "duration": obj.get("duration"),
                "participants": {},  # participant_uuid -> participant data
                "hourly_stats": {},  # hour -> stats
                "chat_messages": [],
                "is_active": True
            }

            # We'll handle scheduling in the lifespan function

    async def _schedule_hourly_reports(self, meeting_uuid: str) -> None:
        """Schedule hourly reports for a meeting"""
        try:
            if meeting_uuid not in self.meetings:
                return

            while meeting_uuid in self.meetings and self.meetings[meeting_uuid]["is_active"]:
                now = pendulum.now(self.et_tz)

                # Find the next full hour
                next_hour = now.replace(minute=0, second=0, microsecond=0).add(hours=1)

                # Sleep until the next hour
                seconds_to_wait = (next_hour - now).total_seconds()
                if seconds_to_wait > 0:
                    await asyncio.sleep(seconds_to_wait)

                # Generate hourly report if the meeting is still active
                if meeting_uuid in self.meetings and self.meetings[meeting_uuid]["is_active"]:
                    await self.generate_hourly_report(meeting_uuid)

                # Add a small delay to avoid multiple executions
                await asyncio.sleep(5)
        except Exception as e:
            print(f"Error in hourly reports scheduler: {e}")

    def _handle_meeting_ended(self, data: Dict[str, Any]) -> None:
        """Handle meeting.ended webhook"""
        if "payload" not in data or "object" not in data["payload"]:
            return

        obj = data["payload"]["object"]
        meeting_uuid = obj.get("uuid")
        if not meeting_uuid or meeting_uuid not in self.meetings:
            return

        # Update meeting state
        self.meetings[meeting_uuid]["is_active"] = False
        self.meetings[meeting_uuid]["end_time"] = obj.get("end_time")

        # Generate final hourly report and EOD report if meeting ends
        asyncio.create_task(self.generate_hourly_report(meeting_uuid))
        asyncio.create_task(self.generate_eod_report(meeting_uuid))

    def _handle_participant_joined(self, data: Dict[str, Any]) -> None:
        """Handle meeting.participant_joined webhook"""
        if "payload" not in data or "object" not in data["payload"]:
            return

        obj = data["payload"]["object"]
        meeting_uuid = obj.get("uuid")
        if not meeting_uuid or meeting_uuid not in self.meetings:
            return

        if "participant" not in obj:
            return

        participant = obj["participant"]
        participant_uuid = participant.get("participant_uuid")
        if not participant_uuid:
            return

        join_time = pendulum.parse(participant.get("join_time"))

        # Update or create participant record
        if participant_uuid not in self.meetings[meeting_uuid]["participants"]:
            self.meetings[meeting_uuid]["participants"][participant_uuid] = {
                "user_name": participant.get("user_name", "Unknown"),
                "email": participant.get("email", ""),
                "sessions": [],
                "chat_count": 0,
                "first_join": join_time
            }

        # Add new session
        self.meetings[meeting_uuid]["participants"][participant_uuid]["sessions"].append({
            "join_time": join_time,
            "leave_time": None
        })

        # Update hourly stats
        hour_key = join_time.in_timezone(self.et_tz).format('YYYY-MM-DD HH:00')
        if hour_key not in self.meetings[meeting_uuid]["hourly_stats"]:
            self.meetings[meeting_uuid]["hourly_stats"][hour_key] = {
                "participants_count": 0,
                "new_participants": 0,
                "chat_count": 0
            }

        self.meetings[meeting_uuid]["hourly_stats"][hour_key]["participants_count"] += 1

        # Check if this is a new participant for this hour
        if len(self.meetings[meeting_uuid]["participants"][participant_uuid]["sessions"]) == 1:
            self.meetings[meeting_uuid]["hourly_stats"][hour_key]["new_participants"] += 1

    def _handle_participant_left(self, data: Dict[str, Any]) -> None:
        """Handle meeting.participant_left webhook"""
        if "payload" not in data or "object" not in data["payload"]:
            return

        obj = data["payload"]["object"]
        meeting_uuid = obj.get("uuid")
        if not meeting_uuid or meeting_uuid not in self.meetings:
            return

        if "participant" not in obj:
            return

        participant = obj["participant"]
        participant_uuid = participant.get("participant_uuid")
        if not participant_uuid or participant_uuid not in self.meetings[meeting_uuid]["participants"]:
            return

        leave_time = pendulum.parse(participant.get("leave_time"))

        # Update the most recent session with leave time
        sessions = self.meetings[meeting_uuid]["participants"][participant_uuid]["sessions"]
        if sessions:
            for session in reversed(sessions):
                if session["leave_time"] is None:
                    session["leave_time"] = leave_time
                    break

    def _handle_chat_message(self, data: Dict[str, Any]) -> None:
        """Handle meeting.chat_message_sent webhook"""
        if "payload" not in data or "object" not in data["payload"]:
            return

        obj = data["payload"]["object"]
        meeting_uuid = obj.get("uuid")
        if not meeting_uuid or meeting_uuid not in self.meetings:
            return

        if "chat_message" not in obj:
            return

        chat = obj["chat_message"]
        sender_name = chat.get("sender_name", "Unknown")
        sender_email = chat.get("sender_email", "")
        message_time = pendulum.parse(chat.get("date_time"))

        # Add to chat messages
        self.meetings[meeting_uuid]["chat_messages"].append({
            "sender_name": sender_name,
            "sender_email": sender_email,
            "time": message_time,
            "content": chat.get("message_content", "")
        })

        # Update hourly stats
        hour_key = message_time.in_timezone(self.et_tz).format('YYYY-MM-DD HH:00')
        if hour_key not in self.meetings[meeting_uuid]["hourly_stats"]:
            self.meetings[meeting_uuid]["hourly_stats"][hour_key] = {
                "participants_count": 0,
                "new_participants": 0,
                "chat_count": 0
            }

        self.meetings[meeting_uuid]["hourly_stats"][hour_key]["chat_count"] += 1

        # Update participant chat count if we can find them
        for participant_uuid, participant_data in self.meetings[meeting_uuid]["participants"].items():
            if (participant_data["user_name"] == sender_name or
                (sender_email and participant_data["email"] == sender_email)):
                participant_data["chat_count"] += 1
                break

    async def generate_hourly_report(self, meeting_uuid: str) -> None:
        """Generate hourly report for a meeting"""
        if meeting_uuid not in self.meetings:
            return

        meeting = self.meetings[meeting_uuid]
        topic = meeting["topic"]

        # Get current time in ET
        now = pendulum.now(self.et_tz)
        date_str = now.format('YYYY-MM-DD')

        # Create directory for reports
        report_dir = self.reports_dir / date_str
        report_dir.mkdir(parents=True, exist_ok=True)

        # Clean topic name - allow only alphanumeric, spaces, and common safe characters
        safe_topic = ''.join(c for c in topic if c.isalnum() or c in ' _-').strip()
        safe_topic = safe_topic.replace(' ', '_')

        # Find a unique filename using counters if needed
        base_filename = f"{safe_topic}_HourlyReport"
        counter = 0
        report_filename = f"{base_filename}.xlsx"
        report_path = report_dir / report_filename

        # If file exists, add counter until we find a unique name
        while report_path.exists():
            counter += 1
            report_filename = f"{base_filename}_{counter:02d}.xlsx"
            report_path = report_dir / report_filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hourly Report"

        # Add header row
        headers = ["Time", "Total Participants", "New Participants Joined", "Chat Interactions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Add data rows
        # Start time for report data (7am ET)
        day_start = now.replace(hour=7, minute=0, second=0, microsecond=0)
        if now.hour < 7:
            day_start = day_start.subtract(days=1)

        # End time for report data (11pm ET or current time, whichever is earlier)
        day_end = day_start.replace(hour=23, minute=0, second=0, microsecond=0)
        if now < day_end:
            day_end = now

        # Convert hourly stats to list of hours
        hourly_data = []
        current = day_start

        # Build list of hours to report
        while current <= day_end:
            hour_key = current.format('YYYY-MM-DD HH:00')
            stats = meeting["hourly_stats"].get(hour_key, {
                "participants_count": 0,
                "new_participants": 0,
                "chat_count": 0
            })

            hourly_data.append({
                "time": current,
                "stats": stats
            })

            current = current.add(hours=1)

        # Fill data
        for row, hour_data in enumerate(hourly_data, 2):
            time_str = hour_data["time"].format('hh:00 A')

            ws.cell(row=row, column=1, value=f"{time_str} ET")
            ws.cell(row=row, column=2, value=hour_data["stats"]["participants_count"])
            ws.cell(row=row, column=3, value=hour_data["stats"]["new_participants"])
            ws.cell(row=row, column=4, value=hour_data["stats"]["chat_count"])

        # Save workbook
        wb.save(report_path)
        print(f"Generated hourly report: {report_path}")

    async def generate_eod_report(self, meeting_uuid: str) -> None:
        """Generate end-of-day report for a meeting"""
        if meeting_uuid not in self.meetings:
            return

        meeting = self.meetings[meeting_uuid]
        topic = meeting["topic"]

        # Get current date in ET
        now = pendulum.now(self.et_tz)
        date_str = now.format('YYYY-MM-DD')

        # Create directory for reports
        report_dir = self.reports_dir / date_str
        report_dir.mkdir(parents=True, exist_ok=True)

        # Clean topic name - allow only alphanumeric, spaces, and common safe characters
        safe_topic = ''.join(c for c in topic if c.isalnum() or c in ' _-').strip()
        safe_topic = safe_topic.replace(' ', '_')

        # Find a unique filename using counters if needed
        base_filename = f"{safe_topic}_EoD"
        counter = 0
        report_filename = f"{base_filename}.xlsx"
        report_path = report_dir / report_filename

        # If file exists, add counter until we find a unique name
        while report_path.exists():
            counter += 1
            report_filename = f"{base_filename}_{counter:02d}.xlsx"
            report_path = report_dir / report_filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EOD Report"

        # Add header row
        headers = ["Participant Name", "Hours Attended", "Percent attended", "Chat Interactions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Day start and end (7am to 11pm ET)
        day_start = now.replace(hour=7, minute=0, second=0, microsecond=0)
        if now.hour < 7:
            day_start = day_start.subtract(days=1)
        day_end = day_start.replace(hour=23, minute=0, second=0, microsecond=0)

        # Total seconds in the day (16 hours)
        total_day_seconds = 16 * 3600

        # Process participant data
        row = 2
        for participant_uuid, participant in meeting["participants"].items():
            total_seconds = 0

            # Calculate total time attended
            for session in participant["sessions"]:
                join_time = session["join_time"]
                leave_time = session["leave_time"] or pendulum.now()

                # Adjust times to be within 7am-11pm ET
                join_time_et = join_time.in_timezone(self.et_tz)
                leave_time_et = leave_time.in_timezone(self.et_tz)

                # Skip sessions outside the 7am-11pm window
                if leave_time_et < day_start or join_time_et > day_end:
                    continue

                # Adjust join and leave times to be within the day window
                if join_time_et < day_start:
                    join_time_et = day_start
                if leave_time_et > day_end:
                    leave_time_et = day_end

                # Add time for this session
                session_seconds = (leave_time_et - join_time_et).total_seconds()
                total_seconds += session_seconds

            # Format hours attended
            hours = int(total_seconds / 3600)
            minutes = int((total_seconds % 3600) / 60)
            hours_attended = f"{hours:02d}:{minutes:02d}"

            # Calculate percentage
            percent_attended = (total_seconds / total_day_seconds) * 100

            # Add row to worksheet
            ws.cell(row=row, column=1, value=participant["user_name"])
            ws.cell(row=row, column=2, value=hours_attended)
            ws.cell(row=row, column=3, value=f"{percent_attended:.1f}%")
            ws.cell(row=row, column=4, value=participant["chat_count"])

            row += 1

        # Save workbook
        wb.save(report_path)
        print(f"Generated EOD report: {report_path}")

# Initialize the meeting state manager
meeting_state = MeetingStateManager()

# Helper function to schedule hourly reports for an active meeting
async def start_hourly_reports_for_meeting(meeting_uuid):
    if meeting_uuid in meeting_state.meetings and meeting_state.meetings[meeting_uuid]["is_active"]:
        asyncio.create_task(meeting_state._schedule_hourly_reports(meeting_uuid))

@app.post("/zoom/webhook")
async def zoom_webhook(request: Request):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{current_time}] Webhook received. Processing...")

    # Get the raw body content
    body = await request.body()
    body_str = body.decode('utf-8')

    # Save raw webhook immediately, before any processing
    try:
        data = json.loads(body_str)
        event_type = data.get("event", "unknown")
        print(f"[{current_time}] Webhook event type: {event_type}")

        # Extract meeting UUID if available
        meeting_uuid = None
        if "payload" in data and "object" in data["payload"]:
            obj = data["payload"]["object"]
            if "uuid" in obj:
                meeting_uuid = obj["uuid"]

        # If we have a meeting UUID, save to the proper structure immediately
        if meeting_uuid:
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            timestamp = now.strftime("%Y-%m-%d_%H-%M-%S-%f")

            # Create directory structure
            raw_dir = pathlib.Path("Raw")
            day_dir = raw_dir / date_str
            meeting_dir = day_dir / meeting_uuid
            meeting_dir.mkdir(parents=True, exist_ok=True)

            # Write webhook data to file immediately
            file_path = meeting_dir / f"{timestamp}.json"
            with open(file_path, 'w') as f:
                json.dump(data, f, indent=2)

            print(f"[{current_time}] Raw webhook saved to {file_path}")
        else:
            # No meeting UUID, save in a general directory
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            timestamp = now.strftime("%Y-%m-%d_%H-%M-%S-%f")

            # Create directory structure
            raw_dir = pathlib.Path("Raw")
            day_dir = raw_dir / date_str
            general_dir = day_dir / "general"
            general_dir.mkdir(parents=True, exist_ok=True)

            # Write webhook data to file
            file_path = general_dir / f"{timestamp}_{event_type}.json"
            with open(file_path, 'w') as f:
                json.dump(data, f, indent=2)

            print(f"[{current_time}] Raw webhook (no meeting UUID) saved to {file_path}")
    except json.JSONDecodeError:
        print(f"[{current_time}] Error parsing webhook JSON")

        # Save invalid JSON to a separate directory
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S-%f")

        raw_dir = pathlib.Path("Raw")
        day_dir = raw_dir / date_str
        error_dir = day_dir / "invalid_json"
        error_dir.mkdir(parents=True, exist_ok=True)

        # Save the raw body that couldn't be parsed
        file_path = error_dir / f"{timestamp}_invalid.txt"
        with open(file_path, 'w') as f:
            f.write(body_str)

        print(f"[{current_time}] Invalid JSON saved to {file_path}")
        raise HTTPException(status_code=400, detail="Invalid JSON")

    # Case 1: Endpoint URL Validation
    if event_type == "endpoint.url_validation":
        print(f"[{current_time}] Processing endpoint validation")
        plain_token = data.get("payload", {}).get("plainToken")
        if not plain_token:
            print(f"[{current_time}] Error: No plain token provided")
            raise HTTPException(status_code=400, detail="No plain token provided")

        # Use the first available token for validation
        current_token = account_manager.get_next_unverified_token()
        encrypted_token = generate_hash(plain_token, current_token)

        # Mark this token as verified
        account_manager.mark_token_as_verified(current_token)

        # Log verification status
        verified_count = sum(account_manager.verified_tokens.values())
        total_count = len(account_manager.tokens)
        print(f"[{current_time}] Verified {verified_count} out of {total_count} accounts")

        print(f"[{current_time}] Validation successful, returning encrypted token")
        return JSONResponse(content={
            "plainToken": plain_token,
            "encryptedToken": encrypted_token
        })

    # Case 2: Regular Webhook Events (no signature verification needed)
    print(f"[{current_time}] Processing regular webhook: {event_type}")

    # Process webhook data
    meeting_state.process_webhook(data)

    # If this is a meeting.started event, schedule hourly reports
    if event_type == "meeting.started" and "payload" in data and "object" in data["payload"]:
        meeting_uuid = data["payload"]["object"].get("uuid")
        if meeting_uuid:
            print(f"[{current_time}] Starting hourly reports for meeting: {meeting_uuid}")
            # Start background task for hourly reports
            asyncio.create_task(start_hourly_reports_for_meeting(meeting_uuid))

    print(f"[{current_time}] Webhook processing complete for {event_type}")
    return {
        "status": "success",
        "message": f"Webhook processed for event: {event_type}"
    }

@app.get("/verification-status")
async def get_verification_status():
    """Endpoint to check verification status of all accounts"""
    verified_count = sum(account_manager.verified_tokens.values())
    total_count = len(account_manager.tokens)
    return {
        "total_accounts": total_count,
        "verified_accounts": verified_count,
        "status_by_token": {
            f"account_{i+1}": verified
            for i, verified in enumerate(account_manager.verified_tokens.values())
        }
    }

@app.get("/test")
async def health_check():
    """Simple health check endpoint"""
    return {"status": "ok"}

@app.post("/reset-token")
async def reset_token(request: Request):
    """Reset verification status for all tokens"""
    try:
        # Optional: If body is provided, reset specific token
        body = await request.body()
        if body:
            try:
                data = json.loads(body)
                token = data.get("token")

                if token:
                    # Check if token exists in our list
                    if token not in account_manager.tokens:
                        raise HTTPException(status_code=404, detail="Token not found")

                    # Reset the verification status for specific token
                    account_manager.verified_tokens[token] = False

                    # Find the account number for this token
                    account_number = account_manager.tokens.index(token) + 1

                    # Save the updated verification status
                    account_manager.save_verification_status()

                    return {
                        "status": "success",
                        "message": f"Verification status reset for account {account_number}",
                        "account_number": account_number
                    }
            except json.JSONDecodeError:
                # If JSON is invalid, continue to reset all tokens
                pass

        # Reset all tokens
        for token in account_manager.tokens:
            account_manager.verified_tokens[token] = False

        # Save the updated verification status
        account_manager.save_verification_status()

        return {
            "status": "success",
            "message": f"Verification status reset for all {len(account_manager.tokens)} accounts",
            "total_accounts": len(account_manager.tokens)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error resetting tokens: {str(e)}")


@app.get("/meetings")
async def get_meetings():
    """Get list of active meetings"""
    active_meetings = {}
    for uuid, meeting in meeting_state.meetings.items():
        if meeting["is_active"]:
            active_meetings[uuid] = {
                "topic": meeting["topic"],
                "start_time": meeting["start_time"],
                "participant_count": len(meeting["participants"])
            }

    return {
        "active_meetings": active_meetings,
        "total_count": len(active_meetings)
    }

@app.get("/generate-reports/{meeting_uuid}")
async def generate_reports(meeting_uuid: str):
    """Manually trigger report generation for a meeting"""
    if meeting_uuid not in meeting_state.meetings:
        raise HTTPException(status_code=404, detail="Meeting not found")

    await meeting_state.generate_hourly_report(meeting_uuid)
    await meeting_state.generate_eod_report(meeting_uuid)

    return {
        "status": "success",
        "message": f"Reports generated for meeting: {meeting_uuid}"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)